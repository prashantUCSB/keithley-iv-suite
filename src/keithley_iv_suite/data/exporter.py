"""CSV and Excel export for all measurement result types."""
from __future__ import annotations

import csv
import dataclasses
import threading
from datetime import datetime
from pathlib import Path
from typing import Union

import numpy as np

from ..measurements.sweep_config import (
    SweepConfig,
    MeasurementType,
)

# Shared lock: callers that export to the *same* Excel workbook from multiple
# threads must pass this (or their own) lock to append_excel_sheet().
_EXCEL_LOCK = threading.Lock()


def default_filename(config: SweepConfig, output_dir: str) -> Path:
    """Return a timestamped default CSV path for *config* inside *output_dir*.

    Example: ``~/Documents/IV_Data/20260410_143522_RESISTOR_IV_run1.csv``
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Use the enum name — already filesystem-safe (e.g. NMOS_TRANSFER)
    mtype_safe = config.measurement_type.name
    if config.label:
        # Keep only alphanumeric, dash, and underscore from the label
        safe_label = "".join(
            c if c.isalnum() or c in ("-", "_") else "_"
            for c in config.label
        ).strip("_")
        filename = f"{ts}_{mtype_safe}_{safe_label}.csv"
    else:
        filename = f"{ts}_{mtype_safe}.csv"
    return Path(output_dir) / filename


def export_csv(
    result: dict,
    path: Union[str, Path],
    config: SweepConfig,
) -> None:
    """Write *result* to a CSV file at *path*.

    The file begins with ``#``-prefixed metadata lines (measurement type,
    timestamp, config parameters) followed by a column-header row and the
    numeric data.

    Parameters
    ----------
    result  : dict returned by the measurement function
    path    : destination file path (created with parents as needed)
    config  : SweepConfig that produced the measurement
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    mtype = config.measurement_type

    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        _write_header(writer, config)

        if mtype == MeasurementType.NMOS_TRANSFER:
            _write_transfer(writer, result)
        elif mtype == MeasurementType.NMOS_OUTPUT:
            _write_output(writer, result)
        elif mtype == MeasurementType.RESISTOR_IV:
            _write_resistor(writer, result)
        elif mtype == MeasurementType.VAN_DER_PAUW:
            _write_vdp(writer, result)
        elif mtype == MeasurementType.HALL_BAR:
            _write_hall(writer, result)
        elif mtype == MeasurementType.FOUR_POINT_PROBE:
            _write_4pp(writer, result)
        elif mtype == MeasurementType.GENERIC_4PORT:
            _write_generic4(writer, result)
        elif mtype == MeasurementType.PHOTODIODE_IV:
            _write_photodiode(writer, result)
        else:
            _write_generic(writer, result)


# ── Private helpers ──────────────────────────────────────────────────────────

def _fmt(v: float) -> str:
    return f"{v:.9g}"


def _write_header(writer: csv.writer, config: SweepConfig) -> None:
    """Emit #-prefixed metadata lines at the top of the file."""
    writer.writerow([f"# Keithley IV Suite — {config.measurement_type.value}"])
    writer.writerow([f"# Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    if config.label:
        writer.writerow([f"# Label: {config.label}"])
    for f in dataclasses.fields(config):
        if f.name in ("measurement_type", "assignments"):
            continue
        writer.writerow([f"# {f.name}: {getattr(config, f.name)}"])
    writer.writerow(["#"])


def _write_transfer(writer: csv.writer, result: dict) -> None:
    vgs    = result["vgs"]
    vgs_s  = result.get("vgs_sensed", np.full_like(vgs, float("nan")))
    id_    = result["id"]
    ig     = result["ig"]
    vds    = result["vds_actual"]
    writer.writerow(["vgs_V", "vgs_sensed_V", "id_A", "ig_A", "vds_actual_V"])
    for row in zip(vgs, vgs_s, id_, ig, vds):
        writer.writerow([_fmt(v) for v in row])


def _write_output(writer: csv.writer, result: dict) -> None:
    writer.writerow(["curve_index", "vgs_V", "vds_forced_V", "vds_sensed_V", "id_A", "ig_A"])
    for ci, curve in enumerate(result["curves"]):
        vgs   = curve["vgs"]
        vds_f = curve["vds_forced"]
        vds_s = curve["vds"]
        id_   = curve["id"]
        ig    = curve["ig"]
        for vds_fi, vds_si, id_i, ig_i in zip(vds_f, vds_s, id_, ig):
            writer.writerow([ci, _fmt(vgs), _fmt(vds_fi), _fmt(vds_si), _fmt(id_i), _fmt(ig_i)])


def _write_resistor(writer: csv.writer, result: dict) -> None:
    vf = result["voltage"]
    vs = result.get("voltage_sensed", np.full_like(vf, float("nan")))
    i_ = result["current"]
    r_ = result["resistance"]
    writer.writerow(["voltage_forced_V", "voltage_sensed_V", "current_A", "resistance_ohm"])
    for row in zip(vf, vs, i_, r_):
        writer.writerow([_fmt(v) for v in row])


def _write_vdp(writer: csv.writer, result: dict) -> None:
    r_fit = result.get("R_fit", float("nan"))
    writer.writerow([f"# R_fit_ohm: {r_fit:.6g}"])
    i_ = result["i_forced"]
    v_ = result["v_sense"]
    r_ = result["resistance_per_pt"]
    writer.writerow(["i_forced_A", "v_sense_V", "resistance_per_pt_ohm"])
    for row in zip(i_, v_, r_):
        writer.writerow([_fmt(v) for v in row])


def _write_hall(writer: csv.writer, result: dict) -> None:
    for key in ("R_xx", "R_xy", "n_sheet", "mu_H", "carrier_type"):
        if key in result:
            writer.writerow([f"# {key}: {result[key]}"])
    i_  = result["i_forced"]
    vl  = result["v_long"]
    vh  = result["v_hall"]
    writer.writerow(["i_forced_A", "v_long_V", "v_hall_V"])
    for row in zip(i_, vl, vh):
        writer.writerow([_fmt(v) for v in row])


def _write_4pp(writer: csv.writer, result: dict) -> None:
    for key in ("R_transfer", "Rs", "rho", "R_sq"):
        if key in result:
            val = result[key]
            writer.writerow([f"# {key}: {val:.6g}" if isinstance(val, float) else f"# {key}: {val}"])
    i_forced = result["i_forced"]
    i_actual = result.get("i_actual", np.full_like(i_forced, float("nan")))
    v_ = result["v_sense"]
    writer.writerow(["i_forced_A", "i_actual_A", "v_sense_V"])
    for row in zip(i_forced, i_actual, v_):
        writer.writerow([_fmt(v) for v in row])


def _write_generic4(writer: csv.writer, result: dict) -> None:
    vf = result["v_forced"]
    vs = result["v_sensed"]
    i_ = result["current"]
    writer.writerow(["v_forced_V", "v_sensed_V", "current_A"])
    for row in zip(vf, vs, i_):
        writer.writerow([_fmt(v) for v in row])


def _write_photodiode(writer: csv.writer, result: dict) -> None:
    vf = result["voltage"]
    vs = result.get("voltage_sensed", np.full_like(vf, float("nan")))
    i_ = result["current"]
    writer.writerow(["voltage_forced_V", "voltage_sensed_V", "current_A"])
    for row in zip(vf, vs, i_):
        writer.writerow([_fmt(v) for v in row])


def _write_generic(writer: csv.writer, result: dict) -> None:
    """Fallback: write all numpy arrays in the result side by side."""
    arrays = {k: v for k, v in result.items() if isinstance(v, np.ndarray)}
    if not arrays:
        return
    writer.writerow(list(arrays.keys()))
    for row in zip(*arrays.values()):
        writer.writerow([_fmt(v) for v in row])


# ── Excel export ─────────────────────────────────────────────────────────────

def append_excel_sheet(
    result: dict,
    config: SweepConfig,
    workbook_path: Union[str, Path],
    lock: threading.Lock = _EXCEL_LOCK,
) -> None:
    """Append one sheet for *result* to an openpyxl workbook at *workbook_path*.

    Creates the file if it does not exist.  Thread-safe when multiple export
    workers share the same *lock* instance.

    Parameters
    ----------
    result        : dict returned by the measurement function
    config        : SweepConfig that produced the measurement
    workbook_path : destination .xlsx file path
    lock          : threading.Lock shared by all concurrent callers for this file
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    path = Path(workbook_path)
    with lock:
        if path.exists():
            wb = load_workbook(path)
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            # Remove the default blank sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Build a unique sheet name (Excel limit: 31 chars)
        base = (config.label or config.measurement_type.name)[:24]
        # Avoid duplicate sheet names
        existing = set(wb.sheetnames)
        sheet_name = base[:31]
        suffix = 1
        while sheet_name in existing:
            sheet_name = f"{base[:28]}_{suffix}"
            suffix += 1

        ws = wb.create_sheet(title=sheet_name)

        # ── Header: dark row ──────────────────────────────────────────────
        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="1F2937")
        hdr_align = Alignment(horizontal="left")

        def _hdr_row(*cells):
            row_idx = ws.max_row + 1
            for col, val in enumerate(cells, start=1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = hdr_align

        _hdr_row("Keithley IV Suite", config.measurement_type.value)
        _hdr_row("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        if config.label:
            _hdr_row("Label", config.label)
        ws.append([])  # blank spacer

        # ── Config parameters ─────────────────────────────────────────────
        _hdr_row("Parameter", "Value")
        for f in dataclasses.fields(config):
            if f.name in ("measurement_type", "assignments"):
                continue
            ws.append([f.name, str(getattr(config, f.name))])
        ws.append([])  # blank spacer

        # ── Data ──────────────────────────────────────────────────────────
        mtype = config.measurement_type
        if mtype == MeasurementType.NMOS_TRANSFER:
            _xlsx_transfer(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.NMOS_OUTPUT:
            _xlsx_output(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.RESISTOR_IV:
            _xlsx_resistor(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.VAN_DER_PAUW:
            _xlsx_vdp(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.HALL_BAR:
            _xlsx_hall(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.FOUR_POINT_PROBE:
            _xlsx_4pp(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.GENERIC_4PORT:
            _xlsx_generic4(ws, result, hdr_font, hdr_fill)
        elif mtype == MeasurementType.PHOTODIODE_IV:
            _xlsx_photodiode(ws, result, hdr_font, hdr_fill)
        else:
            _xlsx_generic(ws, result, hdr_font, hdr_fill)

        # ── Column widths ─────────────────────────────────────────────────
        for col_cells in ws.columns:
            width = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=8,
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(width + 2, 40)

        wb.save(path)


def _xlsx_col_header(ws, headers, hdr_font, hdr_fill):
    """Write a styled column-header row."""
    row_idx = ws.max_row + 1
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col, value=name)
        c.font = hdr_font
        c.fill = hdr_fill


def _xlsx_transfer(ws, result: dict, hf, fill):
    vgs   = result["vgs"]
    vgs_s = result.get("vgs_sensed", np.full_like(vgs, float("nan")))
    id_   = result["id"]
    ig    = result["ig"]
    vds   = result["vds_actual"]
    _xlsx_col_header(ws, ["vgs_V", "vgs_sensed_V", "id_A", "ig_A", "vds_actual_V"], hf, fill)
    for row in zip(vgs, vgs_s, id_, ig, vds):
        ws.append([float(v) for v in row])


def _xlsx_output(ws, result: dict, hf, fill):
    _xlsx_col_header(ws, ["curve_index", "vgs_V", "vds_forced_V", "vds_sensed_V", "id_A", "ig_A"], hf, fill)
    for ci, curve in enumerate(result["curves"]):
        vgs   = curve["vgs"]
        vds_f = curve["vds_forced"]
        vds_s = curve["vds"]
        id_   = curve["id"]
        ig    = curve["ig"]
        for vds_fi, vds_si, id_i, ig_i in zip(vds_f, vds_s, id_, ig):
            ws.append([ci, float(vgs), float(vds_fi), float(vds_si), float(id_i), float(ig_i)])


def _xlsx_resistor(ws, result: dict, hf, fill):
    vf = result["voltage"]
    vs = result.get("voltage_sensed", np.full_like(vf, float("nan")))
    i_ = result["current"]
    r_ = result["resistance"]
    _xlsx_col_header(ws, ["voltage_forced_V", "voltage_sensed_V", "current_A", "resistance_ohm"], hf, fill)
    for row in zip(vf, vs, i_, r_):
        ws.append([float(v) for v in row])


def _xlsx_vdp(ws, result: dict, hf, fill):
    ws.append(["R_fit_ohm", result.get("R_fit", float("nan"))])
    i_ = result["i_forced"]
    v_ = result["v_sense"]
    r_ = result["resistance_per_pt"]
    _xlsx_col_header(ws, ["i_forced_A", "v_sense_V", "resistance_per_pt_ohm"], hf, fill)
    for row in zip(i_, v_, r_):
        ws.append([float(v) for v in row])


def _xlsx_hall(ws, result: dict, hf, fill):
    for key in ("R_xx", "R_xy", "n_sheet", "mu_H", "carrier_type"):
        if key in result:
            ws.append([key, result[key]])
    i_  = result["i_forced"]
    vl  = result["v_long"]
    vh  = result["v_hall"]
    _xlsx_col_header(ws, ["i_forced_A", "v_long_V", "v_hall_V"], hf, fill)
    for row in zip(i_, vl, vh):
        ws.append([float(v) for v in row])


def _xlsx_4pp(ws, result: dict, hf, fill):
    for key in ("R_transfer", "Rs", "rho", "R_sq"):
        if key in result:
            ws.append([key, result[key]])
    i_forced = result["i_forced"]
    i_actual = result.get("i_actual", np.full_like(i_forced, float("nan")))
    v_ = result["v_sense"]
    _xlsx_col_header(ws, ["i_forced_A", "i_actual_A", "v_sense_V"], hf, fill)
    for row in zip(i_forced, i_actual, v_):
        ws.append([float(v) for v in row])


def _xlsx_generic4(ws, result: dict, hf, fill):
    vf = result["v_forced"]
    vs = result["v_sensed"]
    i_ = result["current"]
    _xlsx_col_header(ws, ["v_forced_V", "v_sensed_V", "current_A"], hf, fill)
    for row in zip(vf, vs, i_):
        ws.append([float(v) for v in row])


def _xlsx_photodiode(ws, result: dict, hf, fill):
    vf = result["voltage"]
    vs = result.get("voltage_sensed", np.full_like(vf, float("nan")))
    i_ = result["current"]
    _xlsx_col_header(ws, ["voltage_forced_V", "voltage_sensed_V", "current_A"], hf, fill)
    for row in zip(vf, vs, i_):
        ws.append([float(v) for v in row])


def _xlsx_generic(ws, result: dict, hf, fill):
    """Fallback: write all numpy arrays."""
    arrays = {k: v for k, v in result.items() if isinstance(v, np.ndarray)}
    if not arrays:
        return
    _xlsx_col_header(ws, list(arrays.keys()), hf, fill)
    for row in zip(*arrays.values()):
        ws.append([float(v) for v in row])
